import json
import os
import shutil
import sys
from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook, load_workbook

from parser_cerezas import CAMPOS_SALIDA


CARPETA_ENTRADA = os.path.join("Input", "Entrada")
CARPETA_PROCESADO = "Procesado"
CARPETA_SALIDA = "Salida"
ARCHIVO_EXCEL_SALIDA = "salida_cerezas.xlsx"


def ruta_recurso(nombre_archivo: str) -> str:
    """Busca archivos tanto en desarrollo como dentro de PyInstaller."""
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, nombre_archivo)

    return os.path.join(os.path.dirname(os.path.abspath(__file__)), nombre_archivo)


def ruta_base() -> str:
    """Carpeta desde donde se ejecuta el sistema en modo normal o empaquetado."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)

    return os.path.dirname(os.path.abspath(__file__))


def cargar_diccionarios(ruta: str = "") -> Dict:
    if not ruta:
        ruta = ruta_recurso("diccionarios.json")

    with open(ruta, "r", encoding="utf-8") as archivo:
        return json.load(archivo)


def preparar_carpetas() -> Dict[str, str]:
    base = ruta_base()
    rutas = {
        "base": base,
        "entrada": os.path.join(base, CARPETA_ENTRADA),
        "procesado": os.path.join(base, CARPETA_PROCESADO),
        "salida": os.path.join(base, CARPETA_SALIDA),
    }

    for ruta in rutas.values():
        os.makedirs(ruta, exist_ok=True)

    rutas["excel"] = os.path.join(rutas["salida"], ARCHIVO_EXCEL_SALIDA)
    return rutas


def listar_archivos_entrada(ruta_entrada: str) -> List[str]:
    archivos = []
    for nombre in os.listdir(ruta_entrada):
        ruta = os.path.join(ruta_entrada, nombre)
        if os.path.isfile(ruta) and nombre.lower().endswith(".txt"):
            archivos.append(ruta)

    return sorted(archivos)


def leer_archivo_texto(ruta: str) -> str:
    try:
        with open(ruta, "r", encoding="utf-8") as archivo:
            return archivo.read()
    except UnicodeDecodeError:
        with open(ruta, "r", encoding="utf-8-sig") as archivo:
            return archivo.read()


def agregar_registros_excel(registros: List[Dict], ruta_excel: str) -> None:
    """Agrega filas al mismo Excel en cada ejecucion."""
    if os.path.exists(ruta_excel):
        libro = load_workbook(ruta_excel)
        hoja = libro.active
        if hoja.max_row == 0:
            hoja.append(CAMPOS_SALIDA)
    else:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Registros"
        hoja.append(CAMPOS_SALIDA)

    for registro in registros:
        hoja.append([registro.get(campo, "") for campo in CAMPOS_SALIDA])

    ajustar_ancho_columnas(hoja)
    libro.save(ruta_excel)


def ajustar_ancho_columnas(hoja) -> None:
    for columna in hoja.columns:
        letra = columna[0].column_letter
        largo = 10
        for celda in columna:
            valor = "" if celda.value is None else str(celda.value)
            largo = max(largo, min(len(valor), 60))
        hoja.column_dimensions[letra].width = largo + 2


def mover_a_procesado(ruta_archivo: str, ruta_procesado: str) -> str:
    nombre = os.path.basename(ruta_archivo)
    destino = os.path.join(ruta_procesado, nombre)

    if os.path.exists(destino):
        base, extension = os.path.splitext(nombre)
        marca_tiempo = datetime.now().strftime("%Y%m%d_%H%M%S")
        destino = os.path.join(ruta_procesado, f"{base}_{marca_tiempo}{extension}")

    shutil.move(ruta_archivo, destino)
    return destino
